import cv2
import numpy as np
import math
import os
import sys
from openpyxl import Workbook, load_workbook

# --- 設定 ---
DEFAULT_SCALE = 50.0
# ------------

# グローバル変数
points = []
img_display = None
img_clean = None
scale_ratio = 1.0
is_calibrated = False

def get_memo_input_terminal(prompt_text):
    """
    ターミナル（黒い画面）で入力を受け付ける関数
    """
    print(f"\n======== 入力が必要です ========")
    print(f"【{prompt_text}】")
    memo = input(">> メモを入力してEnterキーを押してください: ")
    print("================================")
    return memo if memo else "No Memo"

def save_to_excel(excel_path, file_name, measure_type, memo, value, unit="um"):
    """
    Excelファイルにデータを追記する関数
    """
    SHEET_ALL = "All_Data"
    SHEET_DIST = "Distance_Only"
    SHEET_AREA = "Area_Only"

    if os.path.exists(excel_path):
        try:
            wb = load_workbook(excel_path)
        except Exception as e:
            print(f"Excel読み込みエラー: {e}")
            return
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    # シート作成
    if SHEET_ALL not in wb.sheetnames:
        ws_all = wb.create_sheet(SHEET_ALL)
        ws_all.append(["ファイル名", "測定タイプ", "メモ", "測定値"])
    else:
        ws_all = wb[SHEET_ALL]

    if SHEET_DIST not in wb.sheetnames:
        ws_dist = wb.create_sheet(SHEET_DIST)
        ws_dist.append(["ファイル名", "メモ", "測定値"])
    else:
        ws_dist = wb[SHEET_DIST]

    if SHEET_AREA not in wb.sheetnames:
        ws_area = wb.create_sheet(SHEET_AREA)
        ws_area.append(["ファイル名", "メモ", "測定値"])
    else:
        ws_area = wb[SHEET_AREA]

    # 書き込み
    val_str = f"{value:.2f} {unit}"
    ws_all.append([file_name, measure_type, memo, val_str])

    if measure_type == "Distance":
        ws_dist.append([file_name, memo, val_str])
    elif measure_type == "Area":
        ws_area.append([file_name, memo, val_str])

    try:
        wb.save(excel_path)
        print(f"保存成功: Excelに書き込みました -> {memo}")
    except PermissionError:
        print("エラー: Excelファイルが開かれています。閉じてから再試行してください。")

def main(image_path, known_scale_um, excel_path):
    global img_display, img_clean, points, scale_ratio, is_calibrated

    if not os.path.exists(image_path):
        print(f"エラー: 画像 '{image_path}' が見つかりません。")
        return

    image_filename_only = os.path.basename(image_path)
    img = cv2.imread(image_path)
    img_clean = img.copy()
    img_display = img.copy()

    print("=== 植物画像解析ツール (Mac修正版) ===")
    print(f"画像: {image_filename_only}")
    print(f"保存先: {excel_path}")
    print("------------------------------------------------")
    print("【ステップ1: スケール調整】")
    print(f"スケールバー（{known_scale_um}μm）の両端をクリックしてください。")

    cv2.namedWindow('Analysis Tool', cv2.WINDOW_NORMAL)
    cv2.setMouseCallback('Analysis Tool', click_event)
    # Macでのウィンドウ前面表示対策
    cv2.setWindowProperty('Analysis Tool', cv2.WND_PROP_TOPMOST, 1)
    cv2.imshow('Analysis Tool', img_display)

    while True:
        key = cv2.waitKey(1) & 0xFF

        # --- キャリブレーション ---
        if not is_calibrated and len(points) == 2:
            p1, p2 = points[0], points[1]
            px_dist = math.sqrt((p1[0]-p2[0])**2 + (p1[1]-p2[1])**2)

            if px_dist > 0:
                scale_ratio = known_scale_um / px_dist
                is_calibrated = True

                cv2.line(img_display, p1, p2, (0, 255, 255), 2)
                cv2.putText(img_display, f"Scale Set: {known_scale_um}um", (p1[0], p1[1]-10),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0,255,255), 2)
                cv2.imshow('Analysis Tool', img_display)

                print(f"完了: 1 pixel = {scale_ratio:.4f} μm")
                print("\n【ステップ2: 測定モード】")
                print("  [d]キー : 距離測定")
                print("  [a]キー : 面積測定")
                print("  [r]キー : リセット")
                print("  [q]キー : 終了")
                points = []
            else:
                points = []

        # --- [d] 距離測定 ---
        elif key == ord('d'):
            if len(points) >= 2:
                p_start, p_end = points[-2], points[-1]
                dist_px = math.sqrt((p_start[0]-p_end[0])**2 + (p_start[1]-p_end[1])**2)
                dist_um = dist_px * scale_ratio

                cv2.line(img_display, p_start, p_end, (0, 255, 0), 2)
                cv2.imshow('Analysis Tool', img_display)
                cv2.waitKey(1) # 描画更新待ち

                # ★ターミナルで入力させる
                print("\n>>> ターミナルを見てください <<<")
                memo = get_memo_input_terminal("距離測定のメモ")

                save_to_excel(excel_path, image_filename_only, "Distance", memo, dist_um, "um")

                mid_x, mid_y = (p_start[0] + p_end[0]) // 2, (p_start[1] + p_end[1]) // 2
                label = f"{memo}: {dist_um:.1f} um"
                cv2.putText(img_display, label, (mid_x, mid_y),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 0), 2)

                print(">>> 測定画面に戻ります <<<")
                points = []
                cv2.imshow('Analysis Tool', img_display)

        # --- [a] 面積測定 ---
        elif key == ord('a'):
            if len(points) >= 3:
                pts_array = np.array(points, np.int32).reshape((-1, 1, 2))
                area_px = cv2.contourArea(pts_array)
                area_um = area_px * (scale_ratio ** 2)

                overlay = img_display.copy()
                cv2.fillPoly(overlay, [pts_array], (255, 0, 255))
                cv2.addWeighted(overlay, 0.4, img_display, 0.6, 0, img_display)
                cv2.polylines(img_display, [pts_array], True, (255, 0, 255), 2)
                cv2.imshow('Analysis Tool', img_display)
                cv2.waitKey(1) # 描画更新待ち

                # ★ターミナルで入力させる
                print("\n>>> ターミナルを見てください <<<")
                memo = get_memo_input_terminal("面積測定のメモ")

                save_to_excel(excel_path, image_filename_only, "Area", memo, area_um, "um2")

                M = cv2.moments(pts_array)
                if M['m00'] != 0:
                    cx, cy = int(M['m10']/M['m00']), int(M['m01']/M['m00'])
                    label = f"{memo}: {area_um:.0f} um2"
                    cv2.putText(img_display, label, (cx-40, cy),
                                cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 2)

                print(">>> 測定画面に戻ります <<<")
                points = []
                cv2.imshow('Analysis Tool', img_display)

        elif key == ord('r'):
            img_display = img_clean.copy()
            points = []
            print("画面リセット")
            cv2.imshow('Analysis Tool', img_display)

        elif key == ord('q'):
            break

    cv2.destroyAllWindows()

def click_event(event, x, y, flags, param):
    global points, img_display
    if event == cv2.EVENT_LBUTTONDOWN:
        points.append((x, y))
        cv2.circle(img_display, (x, y), 3, (0, 0, 255), -1)
        if len(points) > 1 and len(points) < 100:
            cv2.line(img_display, points[-2], points[-1], (0, 0, 255), 1)
        cv2.imshow('Analysis Tool', img_display)

if __name__ == "__main__":
    args = sys.argv
    if len(args) < 2:
        print("使い方: python measure_tool_v3.py <画像パス> [スケール値] [Excelファイル名]")
        sys.exit()

    target_file = args[1]

    target_scale = DEFAULT_SCALE
    if len(args) > 2:
        try:
            target_scale = float(args[2])
        except ValueError:
            print("警告: スケール値不正。デフォルトを使用。")

    target_excel = "measure_results.xlsx"
    if len(args) > 3:
        target_excel = args[3]
        if not target_excel.endswith(".xlsx"):
            target_excel += ".xlsx"

    main(target_file, target_scale, target_excel)
